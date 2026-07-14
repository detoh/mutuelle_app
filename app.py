# app.py — Flask + PostgreSQL + Procédures stockées
import os
import io
import uuid
import pandas as pd
from io import BytesIO
from utils.notifications import send_welcome_email
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
import secrets as secrets_module
from dotenv import load_dotenv
load_dotenv()

# Correction encodage PostgreSQL sous Windows (locale French_France.1252)
# Empêche le UnicodeDecodeError 0xe9 de psycopg2
os.environ.setdefault('PGCLIENTENCODING', 'WIN1252')
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, redirect, url_for, flash, request, send_file, current_app, abort, jsonify
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message
from flask_migrate import Migrate
from flask_wtf import CSRFProtect
from werkzeug.utils import secure_filename
from config import Config
from models import (db, Membre, Cotisation, Aide, Evenement, Don, Projet,
                    BilanAnnuel, FamilleMembre, AssistanceEvenement,ParametresMutuelle,
                    calculer_solde_caisse, get_recent_transactions, get_monthly_flow_data,
                    get_bilan_annuel, get_comparaison_annuelle,
                    get_top_cotisants, get_membres_en_retard, get_cotisations_par_membre,
                    get_aides_assistances_par_membre, get_etat_cotisations_membres,
                    cloturer_exercice, rafraichir_dashboard)

from forms import (LoginForm, ChangerMotDePasseForm, ContactForm, CotisationForm, AideForm,RegisterForm,
                   EvenementForm, ProjetForm, AjoutMembreFamilleForm, DeclarationEvenementForm,ParametresMutuelleForm)

# ─── INITIALISATION ──────────────────────────────────────────────────────────
app = Flask(__name__)
app.config.from_object(Config)
csrf = CSRFProtect(app)
mail = Mail(app)
db.init_app(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

ROUTES_AUTORISEES_SANS_CHANGEMENT_MDP = {
    'changer_mot_de_passe_oblige',
    'login',
    'logout',
    'static',
}

@app.before_request
def forcer_changement_mot_de_passe():
    if not current_user.is_authenticated:
        return None
    if not current_user.doit_changer_mot_de_passe:
        return None
    if request.endpoint in ROUTES_AUTORISEES_SANS_CHANGEMENT_MDP:
        return None
    flash('Veuillez d\'abord changer votre mot de passe.', 'warning')
    return redirect(url_for('changer_mot_de_passe_oblige'))

@app.teardown_appcontext
def shutdown_session(exception=None):
    if exception:
        db.session.rollback()
    db.session.remove()

@app.context_processor
def inject_now():
    return {'now': datetime.now}

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Membre, int(user_id))

# ─── DÉCORATEURS ─────────────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Veuillez vous connecter.', 'danger')
            return redirect(url_for('login'))
        if not current_user.is_admin:
            flash('Accès réservé aux administrateurs.', 'danger')
            return redirect(url_for('mon_compte'))
        return f(*args, **kwargs)
    return decorated_function

def comptable_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Veuillez vous connecter.', 'danger')
            return redirect(url_for('login'))
        if not current_user.is_admin and not current_user.is_comptable:
            flash('Accès réservé aux administrateurs et comptables.', 'danger')
            return redirect(url_for('mon_compte'))
        return f(*args, **kwargs)
    return decorated_function

# ─── UTILITAIRES ─────────────────────────────────────────────────────────────
def envoyer_notification_email(destinataire, sujet, corps):
    try:
        msg = Message(sujet, recipients=[destinataire])
        msg.body = corps
        mail.send(msg)
        return True
    except Exception as e:
        current_app.logger.warning(f"Erreur email: {e}")
        return False
from flask_mail import Message

def envoyer_mail_bienvenue(email, prenom, nom, mot_de_passe):
    """
    Envoie un mail de bienvenue au nouveau membre avec son mot de passe temporaire.
    Retourne True si l'envoi réussit, False sinon (on ne bloque jamais la création
    du membre si le mail échoue — l'admin verra un avertissement).
    """
    try:
        msg = Message(
            subject="Bienvenue à la Mutuelle — Vos identifiants de connexion",
            recipients=[email]
        )
        msg.body = f"""Bonjour {prenom} {nom},

Votre compte a été créé sur l'application de la Mutuelle.

Vos identifiants de connexion :
  Email      : {email}
  Mot de passe : {mot_de_passe}

⚠️  Pour des raisons de sécurité, vous devrez changer ce mot de passe
dès votre première connexion.

Connectez-vous ici : {url_for('login', _external=True)}

Cordialement,
L'administration de la Mutuelle
"""
        msg.html = f"""
<p>Bonjour <strong>{prenom} {nom}</strong>,</p>
<p>Votre compte a été créé sur l'application de la Mutuelle.</p>
<table style="border-collapse:collapse; margin:16px 0;">
  <tr>
    <td style="padding:6px 16px 6px 0; font-weight:bold;">Email</td>
    <td style="padding:6px 0;">{email}</td>
  </tr>
  <tr>
    <td style="padding:6px 16px 6px 0; font-weight:bold;">Mot de passe temporaire</td>
    <td style="padding:6px 0; font-family:monospace; font-size:1.1em;">{mot_de_passe}</td>
  </tr>
</table>
<p style="color:#b45309;">
  ⚠️ Vous devrez changer ce mot de passe dès votre première connexion.
</p>
<p>
  <a href="{url_for('login', _external=True)}"
     style="background:#2563eb; color:white; padding:10px 20px;
            border-radius:6px; text-decoration:none;">
    Se connecter
  </a>
</p>
<p>Cordialement,<br>L'administration de la Mutuelle</p>
"""
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Échec envoi mail à {email} : {e}")
        return False
    
def calculer_donnees_rapport(annee):
    """Centralise tous les calculs du rapport financier annuel, utilisés
    à la fois par la page web et les exports PDF/Excel."""
    annee_prec = annee - 1
    bilan, bilan_prec, comparaison, top_cotisants, monthly_flow = {}, {}, {}, [], []

    try:
        bilan = get_bilan_annuel(annee) or {}
    except Exception as e:
        current_app.logger.error(f"get_bilan_annuel({annee}): {e}")
        db.session.rollback()
    try:
        bilan_prec = get_bilan_annuel(annee_prec) or {}
    except Exception as e:
        current_app.logger.error(f"get_bilan_annuel({annee_prec}): {e}")
        db.session.rollback()
    try:
        comparaison = get_comparaison_annuelle(annee) or {}
    except Exception as e:
        current_app.logger.error(f"get_comparaison_annuelle: {e}")
        db.session.rollback()
    try:
        top_cotisants = get_top_cotisants(annee=annee, limit=10) or []
    except Exception as e:
        current_app.logger.error(f"get_top_cotisants: {e}")
        db.session.rollback()
    try:
        monthly_flow = get_monthly_flow_data(months=12) or []
    except Exception as e:
        current_app.logger.error(f"get_monthly_flow_data: {e}")
        db.session.rollback()

    total_recettes          = bilan.get('total_entrees', 0) or 0
    total_depenses          = bilan.get('total_sorties', 0) or 0
    solde_annuel            = bilan.get('solde_net', 0) or 0
    total_cotisations_annee = bilan.get('total_cotisations', 0) or 0
    total_dons_annee        = bilan.get('total_dons', 0) or 0
    total_aides_annee       = bilan.get('total_aides', 0) or 0
    total_events_annee      = bilan.get('total_evenements', 0) or 0
    total_adhesions_annee   = bilan.get('total_adhesions', 0) or 0
    nb_membres_actifs       = bilan.get('nb_membres_actifs', 0) or 0
    nb_nouveaux_membres     = bilan.get('nb_nouveaux_membres', 0) or 0

    rec_prec = bilan_prec.get('total_entrees', 0) or 0
    dep_prec = bilan_prec.get('total_sorties', 0) or 0
    pourcentage_evolution_recettes = round(((total_recettes - rec_prec) / rec_prec) * 100, 1) if rec_prec > 0 else 0.0
    pourcentage_evolution_depenses = round(((total_depenses - dep_prec) / dep_prec) * 100, 1) if dep_prec > 0 else 0.0
    taux_aides = round(total_aides_annee / total_depenses, 2) if total_depenses > 0 else 0.0
    solde_cumule_final = sum(m.get('solde', 0) or 0 for m in monthly_flow) if monthly_flow else solde_annuel

    MOIS_NOMS = ['Janvier','Fevrier','Mars','Avril','Mai','Juin',
                 'Juillet','Aout','Septembre','Octobre','Novembre','Decembre']
    details_mensuels = []
    solde_cumule = 0
    for i, m in enumerate(monthly_flow or []):
        rec, dep = m.get('entrees', 0) or 0, m.get('sorties', 0) or 0
        sol = rec - dep
        solde_cumule += sol
        details_mensuels.append({
            'nom': MOIS_NOMS[i] if i < 12 else f'Mois {i+1}',
            'cotisations': m.get('cotisations', 0) or 0, 'dons': m.get('dons', 0) or 0,
            'total_recettes': rec, 'aides': m.get('aides', 0) or 0,
            'evenements': m.get('evenements', 0) or 0, 'total_depenses': dep,
            'solde_mensuel': sol, 'solde_cumule': solde_cumule,
        })

    return {
        'annee': annee, 'bilan': bilan, 'comparaison': comparaison,
        'monthly_flow': monthly_flow, 'total_recettes': total_recettes,
        'total_depenses': total_depenses, 'solde_annuel': solde_annuel,
        'total_cotisations_annee': total_cotisations_annee, 'total_dons_annee': total_dons_annee,
        'total_aides_annee': total_aides_annee, 'total_events_annee': total_events_annee,
        'total_adhesions_annee': total_adhesions_annee, 'nb_membres_actifs': nb_membres_actifs,
        'nb_nouveaux_membres': nb_nouveaux_membres,
        'pourcentage_evolution_recettes': pourcentage_evolution_recettes,
        'pourcentage_evolution_depenses': pourcentage_evolution_depenses,
        'taux_aides': taux_aides, 'solde_cumule_final': solde_cumule_final,
        'details_mensuels': details_mensuels,
        'top_membres_cotisants': top_cotisants[:5] if top_cotisants else [],
    }
    
# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — AUTHENTIFICATION
# ═════════════════════════════════════════════════════════════════════════════
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if current_user.is_admin or current_user.is_comptable:
            return redirect(url_for('dashboard'))
        return redirect(url_for('mon_compte'))
    form = LoginForm()
    if form.validate_on_submit():
        user = Membre.query.filter_by(email=form.email.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)

            # 🔒 Interception : mot de passe temporaire pas encore changé.
            # On bloque ICI, avant toute autre redirection, même si l'admin
            # ou le comptable a un mot de passe temporaire à changer aussi.
            if user.doit_changer_mot_de_passe:
                flash('Veuillez changer votre mot de passe avant de continuer.', 'warning')
                return redirect(url_for('changer_mot_de_passe_oblige'))

            next_page = request.args.get('next')
            flash('Connexion réussie !', 'success')
            if next_page:
                return redirect(next_page)
            if user.is_admin or user.is_comptable:
                return redirect(url_for('dashboard'))
            return redirect(url_for('mon_compte'))
        flash('Email ou mot de passe incorrect.', 'danger')
    return render_template('auth/login.html', form=form)


@app.route('/changer-mot-de-passe-oblige', methods=['GET', 'POST'])
@login_required
def changer_mot_de_passe_oblige():
    # Si le membre n'a pas (ou plus) besoin de changer son mot de passe,
    # on ne le laisse pas accéder à cette page pour rien.
    if not current_user.doit_changer_mot_de_passe:
        return redirect(url_for('mon_compte'))

    form = ChangerMotDePasseForm()
    if form.validate_on_submit():
        # sp_changer_password (modifiée) met maintenant aussi à jour
        # doit_changer_mot_de_passe = FALSE directement en base — plus besoin
        # d'UPDATE manuel séparé ici.
        current_user.changer_password(form.nouveau_mot_de_passe.data)
        current_user.doit_changer_mot_de_passe = False

        flash('Mot de passe changé avec succès !', 'success')
        if current_user.is_admin or current_user.is_comptable:
            return redirect(url_for('dashboard'))
        return redirect(url_for('mon_compte'))

    return render_template('auth/changer_mot_de_passe_oblige.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Vous avez été déconnecté.', 'info')
    return redirect(url_for('login'))

@app.route('/contact', methods=['GET', 'POST'])
def contact():
    form = ContactForm()
    if form.validate_on_submit():
        admins = Membre.query.filter_by(is_admin=True).all()
        emails_admins = [a.email for a in admins if a.email]

        sujet = f"[Contact site] {form.sujet.data}"
        corps = f"""Nouveau message reçu via le formulaire de contact du site.

    Nom : {form.nom.data}
    Email : {form.email.data}
    Téléphone : {form.telephone.data or 'Non renseigné'}

Message :
{form.message.data}
"""
        erreurs_envoi = 0
        for email_admin in emails_admins:
            try:
                envoyer_notification_email(email_admin, sujet, corps)
            except Exception as e:
                current_app.logger.error(f"Erreur envoi contact à {email_admin}: {e}")
                erreurs_envoi += 1

        if emails_admins and erreurs_envoi < len(emails_admins):
            flash('Votre message a bien été envoyé. Nous vous répondrons rapidement.', 'success')
        else:
            flash("Votre message n'a pas pu être envoyé pour le moment. Veuillez réessayer plus tard.", 'danger')
        return redirect(url_for('contact'))

    return render_template('contact.html', form=form)

# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
@app.route('/dashboard')
@login_required
def dashboard():
    # Un membre simple n'a pas accès au dashboard principal
    if not (current_user.is_admin or current_user.is_comptable):
        return redirect(url_for('mon_compte'))

    total_entrees = total_sorties = solde_caisse = 0
    nb_membres = 0
    recent_transactions = []
    monthly_flow = []

    try:
        total_entrees, total_sorties, solde_caisse = calculer_solde_caisse()
    except Exception as e:
        current_app.logger.warning(f"Solde caisse indisponible: {e}")
        db.session.rollback()

    try:
        nb_membres = Membre.query.filter_by(statut='Actif').count()
    except Exception as e:
        current_app.logger.warning(f"Comptage membres echoue: {e}")
        db.session.rollback()

    try:
        recent_transactions = get_recent_transactions(limit=20)
    except Exception as e:
        current_app.logger.warning(f"Transactions recentes indisponibles: {e}")
        db.session.rollback()

    try:
        monthly_flow = get_monthly_flow_data(months=12)
    except Exception as e:
        current_app.logger.warning(f"Flux mensuels indisponibles: {e}")
        db.session.rollback()

    return render_template('dashboard.html',
                           total_entrees=total_entrees,
                           total_sorties=total_sorties,
                           solde_caisse=solde_caisse,
                           nb_membres=nb_membres,
                           recent_transactions=recent_transactions,
                           monthly_flow=monthly_flow)

# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — MEMBRES & MON COMPTE
# ═════════════════════════════════════════════════════════════════════════════
@app.route('/mon_compte')
@login_required
def mon_compte():
    # [OK] Intégration AjoutMembreFamilleForm
    form_membre = AjoutMembreFamilleForm()
    
    family_members = FamilleMembre.query.filter_by(
        membre_id=current_user.id
    ).order_by(FamilleMembre.lien_parente, FamilleMembre.nom).all()

    # [NOUVEAU] Formulaire de déclaration d'événement (modal #modalDeclarerEvenement)
    form_evenement = DeclarationEvenementForm()
    form_evenement.famille_id.choices = [(0, '— Sélectionner (facultatif) —')] + [
        (m.id, f"{m.prenom} {m.nom} — {m.lien_parente}") for m in family_members
    ]
    
    cotisations = Cotisation.query.filter_by(membre_id=current_user.id)\
                                  .order_by(Cotisation.date_paiement.desc()).all()
    aides = Aide.query.filter_by(membre_id=current_user.id)\
                      .order_by(Aide.date_attribution.desc()).all()
    assistances = AssistanceEvenement.query.filter_by(membre_id=current_user.id)\
                                            .order_by(AssistanceEvenement.demande_le.desc()).all()

    # [OK] CORRECTION : Filtre SQLAlchemy valide
    today = datetime.now().date()
    cotisations_en_retard = Cotisation.query.filter(
        Cotisation.membre_id == current_user.id,
        Cotisation.statut.in_(['imaye', 'impaye', 'En attente']),
        Cotisation.date_echeance < today
    ).count()
    est_a_jour = (cotisations_en_retard == 0)

    return render_template(
        'members/mon_compte.html',
        user=current_user,
        family_members=family_members,
        cotisations=cotisations,
        aides=aides,
        assistances=assistances,
        est_a_jour=est_a_jour,
        form_membre=form_membre,  # [OK] Transmis au template
        form_evenement=form_evenement  # [NOUVEAU] Transmis au template
    )

@app.route('/membres')
@login_required
@admin_required
def liste_membres():
    page = request.args.get('page', 1, type=int)
    per_page = 25
    recherche = request.args.get('recherche', '', type=str).strip()
    filtre_statut = request.args.get('statut', '', type=str)
    filtre_role = request.args.get('role', '', type=str)

    query = Membre.query

    if recherche:
        like = f"%{recherche}%"
        query = query.filter(
            db.or_(
                Membre.nom.ilike(like),
                Membre.prenom.ilike(like),
                Membre.email.ilike(like),
                db.func.concat(Membre.nom, ' ', Membre.prenom).ilike(like),
                db.func.concat(Membre.prenom, ' ', Membre.nom).ilike(like),
            )
        )

    if filtre_statut:
        query = query.filter(Membre.statut == filtre_statut)

    if filtre_role == 'admin':
        query = query.filter(Membre.is_admin == True)
    elif filtre_role == 'comptable':
        query = query.filter(Membre.is_comptable == True)
    elif filtre_role == 'member':
        query = query.filter(Membre.is_admin == False, Membre.is_comptable == False)

    pagination = query.order_by(Membre.id).paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        'members/liste.html',
        membres=pagination.items,
        page=page,
        total_pages=pagination.pages or 1,
        total_membres=pagination.total,
        recherche=recherche,
        filtre_statut=filtre_statut,
        filtre_role=filtre_role
    )

# ═════════════════════════════════════════════════════════════════════════════
# IMPORT DE MEMBRES EN MASSE (Excel/CSV)
# ═════════════════════════════════════════════════════════════════════════════

# Colonnes attendues dans le fichier d'import. Seules nom/prenom/email sont
# obligatoires ; les autres sont facultatives.
COLONNES_IMPORT_MEMBRES = [
    'nom', 'prenom', 'email', 'telephone',
    'ville', 'date_adhesion', 'fonction', 'service', 'emploi'
]
COLONNES_OBLIGATOIRES_IMPORT = ['nom', 'prenom', 'email']


@app.route('/membres/modele-import')
@login_required
@admin_required
def telecharger_modele_import_membres():
    """Génère et renvoie un fichier Excel modèle vide avec les bons en-têtes."""
    df_modele = pd.DataFrame(columns=COLONNES_IMPORT_MEMBRES)
    # Une ligne d'exemple pour guider le remplissage
    df_modele.loc[0] = [
        'Koffi', 'Awa', 'awa.koffi@mail.com', '0707070707',
        'Abidjan', '2026-01-15', 'Comptable', 'Finance', 'Cadre'
    ]
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_modele.to_excel(writer, index=False, sheet_name='Membres')
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name='modele_import_membres.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/membres/importer', methods=['GET', 'POST'])
@login_required
@admin_required
def importer_membres():
    if request.method == 'GET':
        return render_template('members/importer.html')

    fichier = request.files.get('fichier_import')
    if not fichier or not fichier.filename:
        flash('[ATTENTION] Veuillez sélectionner un fichier.', 'danger')
        return redirect(url_for('importer_membres'))

    extension = fichier.filename.rsplit('.', 1)[-1].lower()
    if extension not in ('xlsx', 'xls', 'csv'):
        flash('[ATTENTION] Formats acceptés : .xlsx, .xls, .csv', 'danger')
        return redirect(url_for('importer_membres'))

    # ── 1. Lecture du fichier ──────────────────────────────────────────────
    try:
        if extension == 'csv':
            df = pd.read_csv(fichier, dtype=str, keep_default_na=False)
        else:
            df = pd.read_excel(fichier, dtype=str, keep_default_na=False)
    except Exception as e:
        flash(f'[ERREUR] Fichier illisible : {e}', 'danger')
        return redirect(url_for('importer_membres'))

    # Normaliser les noms de colonnes (espaces, casse) pour tolérer de
    # petites variations dans le fichier fourni par l'utilisateur.
    df.columns = [str(c).strip().lower() for c in df.columns]

    colonnes_manquantes = [c for c in COLONNES_OBLIGATOIRES_IMPORT if c not in df.columns]
    if colonnes_manquantes:
        flash(
            f"[ERREUR] Colonnes obligatoires manquantes dans le fichier : "
            f"{', '.join(colonnes_manquantes)}. Téléchargez le modèle pour le bon format.",
            'danger'
        )
        return redirect(url_for('importer_membres'))

    # ── 2. Traitement ligne par ligne ──────────────────────────────────────
    from werkzeug.security import generate_password_hash
    import secrets as secrets_module

    lignes_ok = []        # [(numero_ligne, nom, prenom, email, mot_de_passe_genere)]
    lignes_erreur = []    # [(numero_ligne, message)]
    emails_vus_dans_fichier = set()

    for idx, row in df.iterrows():
        numero_ligne = idx + 2  # +2 : ligne 1 = en-têtes, idx 0-based

        nom = str(row.get('nom', '')).strip()
        prenom = str(row.get('prenom', '')).strip()
        email = str(row.get('email', '')).strip().lower()

        if not nom or not prenom or not email:
            lignes_erreur.append((numero_ligne, "nom, prénom ou email manquant"))
            continue

        if '@' not in email or '.' not in email.split('@')[-1]:
            lignes_erreur.append((numero_ligne, f"email invalide : {email}"))
            continue

        if email in emails_vus_dans_fichier:
            lignes_erreur.append((numero_ligne, f"email en doublon dans le fichier : {email}"))
            continue
        emails_vus_dans_fichier.add(email)

        if Membre.query.filter_by(email=email).first():
            lignes_erreur.append((numero_ligne, f"email déjà existant en base : {email}"))
            continue

        telephone = str(row.get('telephone', '')).strip()
        fonction = str(row.get('fonction', '')).strip()
        service = str(row.get('service', '')).strip()
        emploi = str(row.get('emploi', '')).strip()
        ville = str(row.get('ville', '')).strip()

        date_adhesion_brute = str(row.get('date_adhesion', '')).strip()
        date_adhesion = datetime.now().strftime('%Y-%m-%d')
        if date_adhesion_brute:
            # Excel/openpyxl renvoie souvent une date sous la forme
            # 'AAAA-MM-JJ 00:00:00' (datetime complet converti en texte).
            # On ne garde que la partie date avant l'éventuel espace.
            date_adhesion_brute_nettoyee = date_adhesion_brute.split(' ')[0]
            parsed = None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    parsed = datetime.strptime(date_adhesion_brute_nettoyee, fmt)
                    break
                except ValueError:
                    continue
            if parsed:
                date_adhesion = parsed.strftime('%Y-%m-%d')
            else:
                lignes_erreur.append((
                    numero_ligne,
                    f"date_adhesion illisible ({date_adhesion_brute}), format attendu AAAA-MM-JJ"
                ))
                continue

        mot_de_passe_genere = secrets_module.token_urlsafe(8)
        pw_hash = generate_password_hash(mot_de_passe_genere)

        try:
            kwargs_creer = dict(
                nom=nom, prenom=prenom, email=email,
                telephone=telephone, fonction=fonction,
                service=service, emploi=emploi,
                date_adhesion=date_adhesion,
                password_hash=pw_hash, is_admin=False
            )
            # ⚠️ 'ville' n'était pas un paramètre connu de Membre.creer() dans
            # les routes déjà vues (ajouter_membre/register). On l'ajoute ici
            # par précaution ; si models.py ne l'accepte pas encore, retirez
            # cette ligne ou ajoutez le paramètre côté Membre.creer().
            if ville:
                kwargs_creer['ville'] = ville

            Membre.creer(**kwargs_creer)
            lignes_ok.append((numero_ligne, nom, prenom, email, mot_de_passe_genere))
        except TypeError as e:
            # Cas probable : Membre.creer() ne connaît pas 'ville' → on retente sans.
            if 'ville' in kwargs_creer:
                try:
                    kwargs_creer.pop('ville')
                    Membre.creer(**kwargs_creer)
                    lignes_ok.append((numero_ligne, nom, prenom, email, mot_de_passe_genere))
                    continue
                except Exception as e2:
                    db.session.rollback()
                    lignes_erreur.append((numero_ligne, f"erreur création : {e2}"))
                    continue
            db.session.rollback()
            lignes_erreur.append((numero_ligne, f"erreur création : {e}"))
        except Exception as e:
            db.session.rollback()
            if 'EMAIL_DEJA_UTILISE' in str(e):
                lignes_erreur.append((numero_ligne, f"email déjà utilisé : {email}"))
            else:
                lignes_erreur.append((numero_ligne, f"erreur création : {e}"))

    
    # ── 3. Envoi des mails pour les membres importés avec succès ───────────
    mails_ok = 0
    mails_ko = 0
    for (numero_ligne, nom, prenom, email, mot_de_passe_genere) in lignes_ok:
        # On récupère l'objet Membre tout juste créé pour le passer à
        # send_welcome_email (qui attend un objet, pas des champs séparés).
        membre_cree = Membre.query.filter_by(email=email).first()
        if membre_cree and send_welcome_email(membre_cree, mot_de_passe_genere):
            mails_ok += 1
        else:
            mails_ko += 1

    # ── 4. Rapport ──────────────────────────────────────────────────────────
    if lignes_ok:
        flash(f"[OK] {len(lignes_ok)} membre(s) importé(s) avec succès.", 'success')
        if mails_ok:
            flash(f"[MAIL] {mails_ok} mail(s) de bienvenue envoyé(s).", 'success')
        if mails_ko:
            flash(
                f"[ATTENTION] {mails_ko} mail(s) n'ont pas pu être envoyés — "
                f"vérifiez la configuration MAIL dans .env.",
                'warning'
            )
    if lignes_erreur:
        flash(f"[ATTENTION] {len(lignes_erreur)} ligne(s) en erreur — voir le détail ci-dessous.", 'warning')

    return render_template(
        'members/importer.html',
        lignes_ok=lignes_ok,
        lignes_erreur=lignes_erreur,
        termine=True
    )


@app.route('/membre/ajouter', methods=['GET', 'POST'])
@login_required
@admin_required
def ajouter_membre():
    form = RegisterForm()
    if request.method == 'POST':
        email = request.form.get('email')
        if Membre.query.filter_by(email=email).first():
            flash('Cet email est déjà utilisé.', 'danger')
            return redirect(url_for('ajouter_membre'))

        # ✅ Génération automatique du mot de passe temporaire
        # (l'admin ne saisit plus de mot de passe lui-même)
        mot_de_passe_genere = secrets_module.token_urlsafe(8)
        pw_hash = generate_password_hash(mot_de_passe_genere)

        try:
            membre = Membre.creer(
                nom=request.form.get('nom'), prenom=request.form.get('prenom'), email=email,
                telephone=request.form.get('telephone', ''), fonction=request.form.get('fonction', ''),
                service=request.form.get('service', ''), emploi=request.form.get('emploi', ''),
                date_adhesion=request.form.get('date_adhesion', datetime.now().strftime('%Y-%m-%d')),
                password_hash=pw_hash
            )

            # ✅ Envoi du mail de bienvenue avec le mot de passe temporaire
            mail_envoye = send_welcome_email(membre, mot_de_passe_genere)
            if mail_envoye:
                flash(f'Membre ajouté avec succès. Un mail de bienvenue a été envoyé à {email}.', 'success')
            else:
                # On ne bloque jamais la création si le mail échoue, mais on
                # prévient l'admin pour qu'il transmette le mot de passe lui-même.
                flash(
                    f"Membre ajouté avec succès, mais l'envoi du mail a échoué. "
                    f"Mot de passe temporaire : {mot_de_passe_genere}",
                    'warning'
                )

            return redirect(url_for('liste_membres'))
        except Exception as e:
            flash(f'Erreur : {e}', 'danger')
    return render_template('members/form.html', form=form, titre="Ajouter Membre",membre=None)

@app.route('/membre/<int:id>')
@login_required
@admin_required
def detail_membre(id):
    membre = Membre.query.get_or_404(id)
    stats = membre.tableau_de_bord()
    cotisations = Cotisation.query.filter_by(membre_id=membre.id).order_by(Cotisation.date_paiement.desc()).all()
    aides = Aide.query.filter_by(membre_id=membre.id).order_by(Aide.date_attribution.desc()).all()
    return render_template('members/detail.html', membre=membre, stats=stats,
                         total_cotise=stats.get('total_cotise', 0),
                         total_aides=stats.get('total_aides_recues', 0),
                         statut_cotisation=stats.get('statut_cotisation', '—'),
                         derniere_cotisation=stats.get('derniere_cotisation'),
                         cotisations=cotisations, aides=aides)

@app.route('/membre/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_membre(id):
    membre = Membre.query.get_or_404(id)
    form = RegisterForm(obj=membre)
    if request.method == 'POST':
        email = request.form.get('email')
        existing = Membre.query.filter_by(email=email).first()
        if existing and existing.id != membre.id:
            flash('Cet email est déjà utilisé.', 'danger')
            return render_template('members/form.html', form=form, titre=f"Modifier {membre.nom} {membre.prenom}")
        try:
            membre.modifier(
                nom=request.form.get('nom'), prenom=request.form.get('prenom'),
                telephone=request.form.get('telephone', ''), ville=request.form.get('ville', ''),
                fonction=request.form.get('fonction', ''), service=request.form.get('service', ''),
                emploi=request.form.get('emploi', ''), statut=request.form.get('statut', membre.statut)
            )
            membre.email = email
            membre.is_admin = request.form.get('is_admin') == 'on'
            membre.is_comptable = request.form.get('is_comptable') == 'on'
            password = request.form.get('password', '')
            if password:
                membre.changer_password(password)
                # ⚠️ sp_changer_password met doit_changer_mot_de_passe à FALSE
                # systématiquement. Ici, c'est l'ADMIN qui impose ce mot de
                # passe (pas le membre lui-même) → on force le membre à le
                # changer dès sa prochaine connexion, par sécurité.
                db.session.execute(
                    text("UPDATE membres SET doit_changer_mot_de_passe = TRUE WHERE id = :id"),
                    {'id': membre.id}
                )
                db.session.commit()
                membre.doit_changer_mot_de_passe = True
            else:
                db.session.commit()
            flash('Membre modifié avec succès.', 'success')
            return redirect(url_for('liste_membres'))
        except Exception as e:
            flash(f'Erreur : {e}', 'danger')
    return render_template('members/form.html', form=form, titre=f"Modifier {membre.nom} {membre.prenom}",membre=membre)


@app.route('/membre/<int:id>/supprimer', methods=['POST'])
@login_required
@admin_required
def supprimer_membre(id):
    membre = Membre.query.get_or_404(id)
    if membre.id == current_user.id:
        flash('Vous ne pouvez pas supprimer votre propre compte.', 'danger')
        return redirect(url_for('liste_membres'))
    try:
        membre.supprimer()
        flash('Membre supprimé avec succès.', 'success')
    except Exception as e:
        flash(f'Erreur : {e}', 'danger')
    return redirect(url_for('liste_membres'))

@app.route('/membre/<int:id>/statut', methods=['POST'])
@login_required
@admin_required
def changer_statut_membre(id):
    membre = Membre.query.get_or_404(id)
    nouveau_statut = 'Suspendu' if membre.statut == 'Actif' else 'Actif'
    membre.modifier(membre.nom, membre.prenom, membre.telephone, membre.ville,
                    membre.fonction, membre.service, membre.emploi, nouveau_statut)
    flash(f'Statut changé à {nouveau_statut}.', 'success')
    return redirect(url_for('liste_membres'))

@app.route('/admin/designer-comptable/<int:id>', methods=['POST'])
@login_required
@admin_required
def designer_comptable(id):
    membre = Membre.query.get_or_404(id)
    membre.toggle_comptable()
    statut = 'comptable' if membre.is_comptable else 'membre'
    flash(f'{membre.nom} {membre.prenom} est maintenant {statut}.', 'success')
    return redirect(url_for('liste_membres'))

@app.route('/membre/<int:id>/upload_photo', methods=['POST'])
@login_required
def upload_photo(id):
    membre = Membre.query.get_or_404(id)
    if not current_user.is_admin and current_user.id != membre.id:
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('mon_compte'))
    if 'photo' not in request.files or request.files['photo'].filename == '':
        flash('Aucun fichier sélectionné.', 'danger')
        return redirect(url_for('detail_membre', id=id))
    file = request.files['photo']
    filename = secure_filename(f"membre_{id}{datetime.now().strftime('%Y%m%d%H%M%S')}{file.filename}")
    upload_dir = os.path.join('static', 'uploads', 'photos')
    os.makedirs(upload_dir, exist_ok=True)
    file.save(os.path.join(upload_dir, filename))
    membre.maj_photo(f'uploads/photos/{filename}')
    flash('Photo téléchargée avec succès.', 'success')
    return redirect(url_for('detail_membre', id=id))

# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — FINANCES (COTISATIONS, AIDES, DONS)
# ═════════════════════════════════════════════════════════════════════════════
@app.route('/cotisation/ajouter', methods=['GET', 'POST'])
@login_required
@admin_required
def ajouter_cotisation():
    form = CotisationForm()
    form.membre_id.choices = [(m.id, f"{m.nom} {m.prenom}") for m in Membre.query.filter_by(statut='Actif').all()]

    parametres = ParametresMutuelle.get()

    # Pré-remplit le montant suggéré uniquement à l'affichage initial du
    # formulaire (GET), pour ne pas écraser une valeur saisie par l'admin.
    if request.method == 'GET' and not form.montant.data:
        form.montant.data = parametres.montant_par_echeance

    if form.validate_on_submit():
        try:
            Cotisation.enregistrer(
                membre_id=form.membre_id.data, type_cotisation=form.type_cotisation.data,
                montant=form.montant.data, date_paiement=form.date_paiement.data
            )
            flash('Cotisation enregistrée avec succès.', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'Erreur : {e}', 'danger')

    return render_template(
        'finance/cotisation_form.html',
        form=form,
        montant_suggere=parametres.montant_par_echeance,
        frequence_cotisation=parametres.frequence_cotisation
    )

@app.route('/cotisation/<int:id>/recu')
@login_required
def telecharger_recu(id):
    cotisation = Cotisation.query.get_or_404(id)
    membre = cotisation.membre
    if not current_user.is_admin and current_user.id != membre.id:
        flash('Accès non autorisé.', 'danger')
        return redirect(url_for('mon_compte'))
    try:
        # Sécuriser les dates potentiellement None
        if cotisation.date_paiement is None:
            cotisation.date_paiement = datetime.now().date()
        if hasattr(cotisation, 'date_echeance') and cotisation.date_echeance is None:
            cotisation.date_echeance = datetime.now().date()
        if hasattr(membre, 'date_adhesion') and membre.date_adhesion is None:
            membre.date_adhesion = datetime.now().date()

        from utils.pdf_generator import generer_recu_cotisation
        chemin = f"static/uploads/recus/recu_{cotisation.id}.pdf"
        os.makedirs(os.path.dirname(chemin), exist_ok=True)
        generer_recu_cotisation(membre, cotisation, chemin)
        return send_file(chemin, as_attachment=True, download_name=f"recu_{cotisation.id}.pdf")
    except Exception as e:
        current_app.logger.error(f"Erreur génération reçu PDF id={id}: {e}")
        flash(f'Erreur PDF: {str(e)}', 'danger')
        return redirect(url_for('mon_compte'))

@app.route('/aide/ajouter', methods=['GET', 'POST'])
@login_required
@comptable_required
def ajouter_aide():
    form = AideForm()
    form.membre_id.choices = [(m.id, f"{m.nom} {m.prenom}") for m in Membre.query.filter_by(statut='Actif').all()]
    if form.validate_on_submit():
        _, _, solde = calculer_solde_caisse()
        if solde < form.montant.data:
            flash(f'Solde insuffisant ! Solde actuel: {solde:.0f} FCFA', 'danger')
            return redirect(url_for('ajouter_aide'))
        try:
            # ⚠️ Aide.attribuer() doit poser le bon flag de validation selon
            # le rôle du valideur (admin vs comptable). Si Aide.attribuer()
            # (dans models.py) ne le fait pas encore, il faudra l'adapter pour
            # accepter ces paramètres et les utiliser, par ex. :
            #   valide_par_admin = current_user.is_admin
            #   valide_par_comptable = current_user.is_comptable
            Aide.attribuer(
                membre_id=form.membre_id.data, type_aide=form.type_aide.data,
                montant=form.montant.data, description=form.description.data,
                valideur_id=current_user.id,
                valide_par_admin=current_user.is_admin,
                valide_par_comptable=current_user.is_comptable
            )
            flash('Aide accordée avec succès.', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'Erreur : {e}', 'danger')
    return render_template('finance/aide_form.html', form=form)

@app.route('/aides')
@login_required
@comptable_required
def liste_aides():
    aides = Aide.query.order_by(Aide.date_attribution.desc()).all()
    return render_template('finance/liste_aides.html', aides=aides)


@app.route('/aide/<int:id>/supprimer', methods=['POST'])
@login_required
@comptable_required
def supprimer_aide(id):
    aide = Aide.query.get_or_404(id)
    try:
        aide.supprimer()
        flash('Aide supprimée avec succès.', 'success')
    except Exception as e:
        flash(f'Erreur : {e}', 'danger')
    return redirect(url_for('liste_aides'))



@app.route('/don/ajouter', methods=['GET', 'POST'])
@login_required
@admin_required
def ajouter_don():
    if request.method == 'POST':
        try:
            # 1. Récupérer le statut et le convertir en booléen
            # Si le select envoie 'true', la condition est Vraie, donc statut_recu = True
            statut_recu = request.form.get('statut') == 'true'
            
            # 2. Passer le statut à la méthode d'enregistrement
            Don.enregistrer(
                donateur_nom=request.form.get('donateur_nom'),
                montant=float(request.form.get('montant')),
                statut=statut_recu  # <-- NOUVEAU PARAMÈTRE
            )
            flash('Don enregistré avec succès.', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            db.session.rollback() # Bonne pratique en cas d'erreur
            flash(f'Erreur lors de l\'enregistrement : {e}', 'danger')
            
    return render_template('finance/don_form.html')


# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — ÉVÉNEMENTS & PROJETS
# ═════════════════════════════════════════════════════════════════════════════
@app.route('/evenements')
@login_required
@admin_required
def liste_evenements():
    evenements = Evenement.query.order_by(Evenement.date.desc()).all()
    total_budget   = sum(e.budget_prevu     for e in evenements) or 0
    total_depenses = sum(e.depenses_reelles for e in evenements) or 0
    return render_template('evenements/liste.html', evenements=evenements,
                         total_budget=total_budget, total_depenses=total_depenses)

@app.route('/evenement/ajouter', methods=['GET', 'POST'])
@login_required
@admin_required
def ajouter_evenement():
    form = EvenementForm()
    if form.validate_on_submit():
        try:
            Evenement.creer(
                nom=form.nom.data, 
                budget_prevu=form.budget_prevu.data,
                depenses_reelles=form.depenses_reelles.data, 
                date=form.date.data,
                description=form.description.data,
                statut=form.statut.data  #  AJOUT ICI (c'est déjà un booléen)
            )
            flash('Événement créé avec succès.', 'success')
            return redirect(url_for('liste_evenements'))
        except Exception as e:
            db.session.rollback() # Toujours bon à ajouter en cas d'erreur pour débloquer la session
            flash(f'Erreur lors de la création : {e}', 'danger')
            
    return render_template('evenements/form.html', form=form, titre="Nouvel Événement")

@app.route('/evenement/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_evenement(id):
    evenement = Evenement.query.get_or_404(id)
    form = EvenementForm(obj=evenement)
    if form.validate_on_submit():
        try:
            evenement.modifier(
                nom=form.nom.data, budget_prevu=form.budget_prevu.data,
                depenses_reelles=form.depenses_reelles.data, date=form.date.data,
                description=form.description.data
            )
            flash('Événement modifié avec succès.', 'success')
            return redirect(url_for('liste_evenements'))
        except Exception as e:
            flash(f'Erreur : {e}', 'danger')
    return render_template('evenements/form.html', form=form, titre=f"Modifier {evenement.nom}")

@app.route('/evenement/<int:id>/supprimer', methods=['POST'])
@login_required
@admin_required
def supprimer_evenement(id):
    evenement = Evenement.query.get_or_404(id)
    try:
        evenement.supprimer()
        flash('Événement supprimé avec succès.', 'success')
    except Exception as e:
        flash(f'Erreur : {e}', 'danger')
    return redirect(url_for('liste_evenements'))

@app.route('/projets')
@login_required
@admin_required
def liste_projets():
    projets = Projet.query.order_by(Projet.date_creation.desc()).all()
    return render_template('projets/liste.html', projets=projets)

@app.route('/projet/ajouter', methods=['GET', 'POST'])
@login_required
@admin_required
def ajouter_projet():
    form = ProjetForm()
    if form.validate_on_submit():
        try:
            Projet.creer(
                titreprojet=form.titreprojet.data, description=form.description.data,
                datedebut=form.datedebut.data, datefiprojet=form.datefiprojet.data,
                coutexecution=form.coutexecution.data, nomdonateur=form.nomdonateur.data,
                statut=form.statut.data
            )
            flash('Projet créé avec succès.', 'success')
            return redirect(url_for('liste_projets'))
        except Exception as e:
            flash(f'Erreur : {e}', 'danger')
    return render_template('projets/form.html', form=form, titre="Nouveau Projet")

@app.route('/projet/<int:id>/modifier', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_projet(id):
    projet = Projet.query.get_or_404(id)
    form = ProjetForm(obj=projet)
    if form.validate_on_submit():
        try:
            projet.modifier(
                titreprojet=form.titreprojet.data, description=form.description.data,
                datedebut=form.datedebut.data, datefiprojet=form.datefiprojet.data,
                coutexecution=form.coutexecution.data, nomdonateur=form.nomdonateur.data,
                statut=form.statut.data
            )
            flash('Projet modifié avec succès.', 'success')
            return redirect(url_for('liste_projets'))
        except Exception as e:
            flash(f'Erreur : {e}', 'danger')
    return render_template('projets/form.html', form=form, titre=f"Modifier {projet.titreprojet}")

@app.route('/projet/<int:id>/supprimer', methods=['POST'])
@login_required
@admin_required
def supprimer_projet(id):
    projet = Projet.query.get_or_404(id)
    try:
        projet.supprimer()
        flash('Projet supprimé avec succès.', 'success')
    except Exception as e:
        flash(f'Erreur : {e}', 'danger')
    return redirect(url_for('liste_projets'))

# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — COMPTABLE & ASSISTANCES FAMILIALES
# ═════════════════════════════════════════════════════════════════════════════
@app.route('/comptable')
@login_required
@comptable_required
def dashboard_comptable():
    from sqlalchemy import func
    total_entrees, total_sorties, solde_caisse = calculer_solde_caisse()
    today = datetime.now()
    mois_debut = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    cotisations_mois = db.session.query(func.sum(Cotisation.montant)).filter(
        Cotisation.date_paiement >= mois_debut, Cotisation.statut == 'Paye'
    ).scalar() or 0
    aides_mois = db.session.query(func.sum(Aide.montant)).filter(
        Aide.date_attribution >= mois_debut
    ).scalar() or 0
    cotisations_impayees  = Cotisation.query.filter_by(statut='En attente').count()
    recent_transactions   = get_recent_transactions(limit=20)
    monthly_flow          = get_monthly_flow_data(months=12)
    #membres_en_retard     = get_membres_en_retard(mois_retard=1)
    membres_en_retard = get_membres_en_retard()

    return render_template('comptable/dashboard.html',
                         solde=solde_caisse, entrees=total_entrees, sorties=total_sorties,
                         cotisations_mois=cotisations_mois, aides_mois=aides_mois,
                         cotisations_impayees=cotisations_impayees,
                         recent_transactions=recent_transactions, monthly_flow=monthly_flow,
                         membres_en_retard=membres_en_retard[:10])

@app.route('/comptable/transactions')
@login_required
@comptable_required
def transactions_comptable():
    page    = request.args.get('page', 1, type=int)
    per_page = 25
    all_transactions  = get_recent_transactions(limit=500)
    start, end        = (page - 1) * per_page, page * per_page
    transactions_page = all_transactions[start:end]
    total_pages       = (len(all_transactions) + per_page - 1) // per_page
    return render_template('comptable/transactions.html',
                         transactions=transactions_page, page=page, total_pages=total_pages)

@app.route('/comptable/valider-cotisation/<int:id>', methods=['POST'])
@login_required
@comptable_required
def valider_cotisation(id):
    cotisation = Cotisation.query.get_or_404(id)
    cotisation.statut = 'Paye'
    db.session.commit()
    flash('Cotisation validée avec succès.', 'success')
    return redirect(url_for('transactions_comptable'))

@app.route('/comptable/rapports')
@login_required
@comptable_required
def rapports_comptable():
    annee = request.args.get('annee', datetime.now().year, type=int)
    date_debut = datetime(annee, 1, 1)
    date_fin   = datetime(annee, 12, 31, 23, 59, 59)
    
    bilan = get_bilan_annuel(annee)
    comparaison = get_comparaison_annuelle(annee)
    top_cotisants = get_top_cotisants(annee=annee, limit=10)
    
    cotisations_par_membre = get_cotisations_par_membre(annee)
    aides_assistances_par_membre = get_aides_assistances_par_membre(annee)
    dons = Don.query.filter(Don.date.between(date_debut, date_fin)).order_by(Don.date.desc()).all()
    bilans_archives = BilanAnnuel.query.order_by(BilanAnnuel.annee.desc()).all()

    return render_template('comptable/rapports.html', annee=annee, bilan=bilan, comparaison=comparaison,
                         top_cotisants=top_cotisants, total_cotisations=bilan.get('total_cotisations', 0),
                         total_aides=bilan.get('total_aides', 0), total_dons=bilan.get('total_dons', 0),
                         total_evenements=bilan.get('total_evenements', 0), total_projets=bilan.get('total_projets', 0),
                         solde_annee=bilan.get('solde_net', 0), cotisations_par_membre=cotisations_par_membre,
                         aides_assistances_par_membre=aides_assistances_par_membre, dons=dons,
                         bilans_archives=bilans_archives)

@app.route('/comptable/etat-cotisations')
@login_required
@comptable_required
def etat_cotisations():
    annee = request.args.get('annee', datetime.now().year, type=int)
    page = request.args.get('page', 1, type=int)
    recherche = request.args.get('recherche', '', type=str).strip()
    filtre_statut = request.args.get('statut', '', type=str)
    par_page = 25

    etat_complet, nb_echeances = get_etat_cotisations_membres(annee)

    # Filtre recherche (nom ou prénom, insensible à la casse)
    if recherche:
        recherche_lower = recherche.lower()
        etat_complet = [
            m for m in etat_complet
            if recherche_lower in m['nom'].lower()
            or recherche_lower in m['prenom'].lower()
            or recherche_lower in f"{m['nom']} {m['prenom']}".lower()
            or recherche_lower in f"{m['prenom']} {m['nom']}".lower()
        ]

    # Filtre statut
    if filtre_statut in ('À jour', 'Pas à jour'):
        etat_complet = [m for m in etat_complet if m['statut'] == filtre_statut]

    total_membres = len(etat_complet)
    total_pages = max(1, (total_membres + par_page - 1) // par_page)
    page = max(1, min(page, total_pages))

    debut = (page - 1) * par_page
    fin = debut + par_page
    etat_page = etat_complet[debut:fin]

    return render_template(
        'comptable/etat_cotisations.html',
        etat=etat_page,
        nb_echeances=nb_echeances,
        annee=annee,
        page=page,
        total_pages=total_pages,
        total_membres=total_membres,
        recherche=recherche,
        filtre_statut=filtre_statut
    )
# ═════════════════════════════════════════════════════════════════════════════
# ROUTE — DÉCLARATION D'ÉVÉNEMENT PAR LE MEMBRE (modal mon_compte.html)
# La méthode du modèle s'appelle creer_demande (voir models.py ligne 243)
# ═════════════════════════════════════════════════════════════════════════════
ALLOWED_EXTENSIONS_EV = {'pdf', 'jpg', 'jpeg', 'png'}
LABELS_EVENEMENT = {
    'Deces': 'Décès',
    'Naissance': 'Naissance',
    'Mariage': 'Mariage',
    'Accident': 'Accident',
    'Retraite': 'Retraite',
    'Mutation': 'Mutation',
    'Autre': 'Autre',
    # 'Maladie' volontairement non activé pour l'instant
}

def _allowed_piece(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_EV

@app.route('/mon_compte/declarer-evenement', methods=['POST'])
@login_required
def declarer_evenement_membre():
    form = DeclarationEvenementForm()

    # famille_id : choices dynamiques (membres de famille du membre connecté).
    # Doit être renseigné avant validate_on_submit() pour que WTForms valide
    # la valeur postée contre cette liste.
    family_members = FamilleMembre.query.filter_by(membre_id=current_user.id).all()
    form.famille_id.choices = [(0, '— Sélectionner (facultatif) —')] + [
        (m.id, f"{m.prenom} {m.nom}") for m in family_members
    ]

    if not form.validate_on_submit():
        for field_errors in form.errors.values():
            for err in field_errors:
                flash(f'[ATTENTION] {err}', 'danger')
        return redirect(url_for('mon_compte'))

    # ── 1. Lecture des champs (désormais validés par WTForms) ─────────────
    type_evt    = form.type_evenement.data
    date_evt    = form.date_evenement.data.strftime('%Y-%m-%d')
    description = (form.description.data or '').strip()
    montant     = float(form.montant_demande.data) if form.montant_demande.data else 0.0
    famille_id  = form.famille_id.data if form.famille_id.data else None
    notes       = (form.notes.data or '').strip() or None

    # ── 2. Validation métier spécifique ────────────────────────────────────
    # ⚠️ La logique ci-dessous (LABELS_EVENEMENT, dossiers d'upload, email)
    # ne couvre actuellement que ces 3 types. Le formulaire en propose 8
    # (Maladie, Accident, Retraite, Mutation, Autre en plus). Pour l'instant
    # on bloque les types non encore supportés côté métier — à étendre dans
    # LABELS_EVENEMENT / dossiers si vous voulez les activer.
    if type_evt not in LABELS_EVENEMENT:
        flash(
            f"[ATTENTION] Le type « {type_evt} » n'est pas encore pris en charge. "
            "Contactez l'administration pour ce type de demande.", 'warning'
        )
        return redirect(url_for('mon_compte'))
    if not description:
        flash('[ATTENTION] La description est obligatoire.', 'danger')
        return redirect(url_for('mon_compte'))

    # ── 3. Types concernant uniquement le membre (pas un proche) ──────────
    # Mariage, Accident, Retraite, Mutation et Autre concernent directement
    # le membre déclarant : on ignore famille_id même s'il a été posté.
    TYPES_MEMBRE_SEUL = {'Mariage', 'Accident', 'Retraite', 'Mutation', 'Autre'}
    if type_evt in TYPES_MEMBRE_SEUL:
        famille_id = None

    # ── 4. Vérifier que famille_id appartient bien au membre ──────────────
    fam = None
    if famille_id:
        fam = FamilleMembre.query.filter_by(
            id=famille_id, membre_id=current_user.id
        ).first()
        if not fam:
            flash('[ATTENTION] Membre de famille introuvable.', 'danger')
            return redirect(url_for('mon_compte'))

    filtre = dict(membre_id=current_user.id, type_evenement=type_evt, statut='En attente')
    if famille_id:
        filtre['famille_id'] = famille_id
    if AssistanceEvenement.query.filter_by(**filtre).first():
        flash(f'[INFO] Une demande « {LABELS_EVENEMENT[type_evt]} » est déjà en attente de validation.', 'info')
        return redirect(url_for('mon_compte'))

    # ── 6. Pièce justificative (facultative) ──────────────────────────────
    piece_path = None
    fichier = form.piece_justificative.data
    if fichier and fichier.filename:
        if not _allowed_piece(fichier.filename):
            flash('[ATTENTION] Format non autorisé. Utilisez PDF, JPG ou PNG.', 'warning')
            return redirect(url_for('mon_compte'))
        dossiers     = {
            'Deces': 'deces',
            'Naissance': 'naissances',
            'Mariage': 'mariages',
            'Accident': 'accidents',
            'Retraite': 'retraites',
            'Mutation': 'mutations',
            'Autre': 'autres',
        }
        safe_name    = secure_filename(fichier.filename)
        sous_dossier = dossiers.get(type_evt, 'assistances')
        upload_dir   = os.path.join(current_app.root_path, 'static', 'uploads', sous_dossier)
        os.makedirs(upload_dir, exist_ok=True)
        nom_fichier  = f"{current_user.id}_{date_evt}_{safe_name}"
        fichier.save(os.path.join(upload_dir, nom_fichier))
        piece_path = f"uploads/{sous_dossier}/{nom_fichier}"

    # ── 7. Enregistrement — méthode creer_demande (models.py l.243) ───────
    try:
        AssistanceEvenement.creer_demande(
            membre_id      = current_user.id,
            type_evenement = type_evt,
            date_evenement = date_evt,
            description    = description,
            montant_prevu  = montant,
            famille_id     = famille_id,
            piece_path     = piece_path,
            notes          = notes
        )

        # ── 8. Notification email admin (non bloquante) ───────────────────
        try:
            admin = Membre.query.filter_by(is_admin=True).first()
            if admin and admin.email:
                proche = f" (concernant {fam.prenom} {fam.nom})" if fam else ""
                envoyer_notification_email(
                    destinataire = admin.email,
                    sujet        = f"[Mutuelle] Nouvelle déclaration : {LABELS_EVENEMENT[type_evt]}",
                    corps        = (
                        f"Le membre {current_user.prenom} {current_user.nom} a soumis "
                        f"une déclaration « {LABELS_EVENEMENT[type_evt]} »{proche}.\n"
                        f"Date : {date_evt}\nDescription : {description}"
                    )
                )
        except Exception as mail_err:
            current_app.logger.warning(f"Email admin non envoyé : {mail_err}")

        flash(
            f'[OK] Déclaration « {LABELS_EVENEMENT[type_evt]} » soumise avec succès. '
            'En attente de validation par l\'administration.',
            'success'
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur déclaration événement membre {current_user.id}: {e}")
        flash(f'[ERREUR] Erreur lors de la déclaration : {e}', 'danger')

    return redirect(url_for('mon_compte'))


@app.route('/admin/assistances/<int:assistance_id>/valider', methods=['POST'])
@login_required
def valider_assistance(assistance_id):
    if not (current_user.is_admin or current_user.is_comptable):
        abort(403)
    assistance = AssistanceEvenement.query.get_or_404(assistance_id)

    if assistance.type_evenement == 'Naissance' and assistance.famille:
        enfant = assistance.famille
        if hasattr(enfant, 'est_enfant_membres') and enfant.est_enfant_membres:
            autre_id = enfant.parent2_membre_id if assistance.membre_id == enfant.parent1_membre_id else enfant.parent1_membre_id
            if autre_id:
                deja_versee = AssistanceEvenement.query.filter_by(
                    type_evenement='Naissance', famille_id=enfant.id,
                    membre_id=autre_id, statut='Versée'
                ).first()
                if deja_versee:
                    flash('[ATTENTION] L\'autre parent a déjà reçu l\'assistance pour cette naissance. Validation refusée.', 'warning')
                    return redirect(url_for('admin_assistances'))

    assistance.valider(valideur_id=current_user.id)
    flash(f'[OK] Assistance de {assistance.montant_prevu:,.0f} FCFA versée', 'success')
    return redirect(url_for('admin_assistances'))

@app.route('/admin/assistances/<int:assistance_id>/refuser', methods=['POST'])
@login_required
def refuser_assistance(assistance_id):
    if not (current_user.is_admin or current_user.is_comptable):
        abort(403)
    assistance = AssistanceEvenement.query.get_or_404(assistance_id)
    motif = request.form.get('motif_refus', '').strip()

    if not motif:
        flash('[ATTENTION] Le motif du refus est obligatoire.', 'warning')
        return redirect(url_for('admin_assistances'))

    assistance.refuser(motif=motif, valideur_id=current_user.id)
    flash(f'[ERREUR] Demande #{assistance.id} refusée. Motif enregistré.', 'info')
    return redirect(url_for('admin_assistances'))

# [OK] INTÉGRATION AJOUTMEMBREFAAMILLEFORM
@app.route('/ajouter_membre_famille', methods=['GET', 'POST'])
@login_required
def ajouter_membre_famille():
    form = AjoutMembreFamilleForm()
    if form.validate_on_submit():
        nouveau_membre = FamilleMembre(
            membre_id=current_user.id,
            nom=form.nom.data, prenom=form.prenom.data,
            lien_parente=form.lien_parente.data,
            date_naissance=form.date_naissance.data,
            telephone=form.telephone.data, adresse=form.adresse.data,
            est_a_charge=form.est_a_charge.data, notes=form.notes.data
        )
        db.session.add(nouveau_membre)
        db.session.commit()
        flash('Membre de famille ajouté avec succès [OK]', 'success')
        return redirect(url_for('mon_compte'))
    return redirect(url_for('mon_compte'))

# ═════════════════════════════════════════════════════════════════════════════
# ROUTES — RAPPORTS & ADMIN DIVERS
# ═════════════════════════════════════════════════════════════════════════════
@app.route('/rapports')
@login_required
@admin_required
def rapports():
    annee = request.args.get('annee', datetime.now().year, type=int)
    d = calculer_donnees_rapport(annee)
    return render_template(
        'rapports/index.html',
        annee_selectionnee=d['annee'],
        years_list=list(range(2020, datetime.now().year + 2)),
        **d
    )

@app.route('/comptable/etat-cotisations/export/pdf')
@login_required
@comptable_required
def export_etat_cotisations_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO

    annee = request.args.get('annee', datetime.now().year, type=int)
    etat, nb_echeances = get_etat_cotisations_membres(annee)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"État des Cotisations — Année {annee}", styles['Title']))
    elements.append(Spacer(1, 0.5*cm))

    # En-têtes dynamiques selon la fréquence
    entetes = ['Membre', 'Adhésion'] + [f'Trim. {i}' for i in range(1, nb_echeances + 1)] + ['Payé', 'Attendu', 'Solde', 'Statut']
    data = [entetes]

    for m in etat:
        ligne = [
            f"{m['nom']} {m['prenom']}",
            str(m['date_adhesion']),
        ]
        ligne += ['OK' if p['payee'] else '-' for p in m['periodes']]
        ligne += [
            f"{m['total_paye']:,.0f}",
            f"{m['total_attendu']:,.0f}",
            f"{m['solde']:,.0f}",
            m['statut'],
        ]
        data.append(ligne)

    table = Table(data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ])

    # Colore en rouge la ligne des membres "Pas à jour" (dernière colonne = Statut)
    for idx, m in enumerate(etat, start=1):
        if m['statut'] == 'Pas à jour':
            style.add('TEXTCOLOR', (-1, idx), (-1, idx), colors.red)

    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'etat_cotisations_{annee}.pdf',
        mimetype='application/pdf'
    )

@app.route('/comptable/etat-cotisations/export/excel')
@login_required
@comptable_required
def export_etat_cotisations_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    annee = request.args.get('annee', datetime.now().year, type=int)
    etat, nb_echeances = get_etat_cotisations_membres(annee)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Cotisations {annee}"

    entetes = ['Membre', 'Adhésion'] + [f'Échéance {i}' for i in range(1, nb_echeances + 1)] + ['Payé', 'Attendu', 'Solde', 'Statut']
    ws.append(entetes)

    en_tete_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    en_tete_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = en_tete_fill
        cell.font = en_tete_font
        cell.alignment = Alignment(horizontal='center')

    for m in etat:
        ligne = [
            f"{m['nom']} {m['prenom']}",
            str(m['date_adhesion']),
        ]
        ligne += ['Oui' if p['payee'] else 'Non' for p in m['periodes']]
        ligne += [m['total_paye'], m['total_attendu'], m['solde'], m['statut']]
        ws.append(ligne)

        # Colore en rouge la ligne si le membre n'est pas à jour
        if m['statut'] == 'Pas à jour':
            for cell in ws[ws.max_row]:
                cell.font = Font(color='C0392B')

    # Ajuste la largeur des colonnes
    for col in ws.columns:
        largeur_max = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(largeur_max + 2, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'etat_cotisations_{annee}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@app.route('/admin/test-email')
@login_required
@admin_required
def test_email():
    from utils.notifications import send_email
    success = send_email(
        subject=" Test Email - Mutuelle Solidaire", recipients=[current_user.email],
        text_body="Ceci est un email de test.",
        html_body="<h1>Test Email [OK]</h1><p>Configuration email fonctionnelle !</p>"
    )
    flash('Email de test envoyé !' if success else 'Échec envoi email.', 'success' if success else 'danger')
    return redirect(url_for('dashboard'))

@app.route('/admin/valider-cotisations', methods=['POST'])
@login_required
@admin_required
def admin_valider_cotisations():
    try:
        # 1. Récupérer toutes les cotisations qui sont en attente
        # (On utilise ilike pour être sûr de capturer 'en_attente', 'En attente', etc.)
        cotisations_a_valider = Cotisation.query.filter(Cotisation.statut.ilike('%en_attente%')).all()
        
        if not cotisations_a_valider:
            flash('Aucune cotisation en attente de validation.', 'info')
            return redirect(url_for('dashboard')) # Adaptez si vous avez une route 'liste_cotisations'

        # 2. Mettre à jour le statut de chaque cotisation
        # On met 'Paye' (sans accent) pour compatibilite WIN1252 et fn_solde_caisse
        for cot in cotisations_a_valider:
            cot.statut = 'Paye'
            
        db.session.commit()
        flash(f'{len(cotisations_a_valider)} cotisation(s) validée(s) avec succès !', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur lors de la validation des cotisations : {e}")
        flash(f'Erreur lors de la validation : {e}', 'danger')
        
    return redirect(url_for('dashboard')) # Adaptez la redirection selon votre besoin

@app.route('/comptable/membres-en-retard')
@login_required
@comptable_required
def membres_en_retard():
    parametres = ParametresMutuelle.get()
    mois_defaut = 12 // parametres.nb_echeances_par_an

    mois = request.args.get('mois', mois_defaut, type=int)
    recherche = request.args.get('recherche', '', type=str).strip()
    page = request.args.get('page', 1, type=int)
    par_page = 25

    retardataires_complet = get_membres_en_retard(mois_retard=mois)
    nb_retard = len(retardataires_complet)  # total réel, indépendant de la recherche

    # Filtre recherche (nom, prénom, ou nom complet — nom_complet contient déjà les deux)
    if recherche:
        recherche_lower = recherche.lower()
        retardataires_filtres = [
            r for r in retardataires_complet
            if recherche_lower in r['nom_complet'].lower()
        ]
    else:
        retardataires_filtres = retardataires_complet

    nb_filtres = len(retardataires_filtres)
    total_pages = max(1, (nb_filtres + par_page - 1) // par_page)
    page = max(1, min(page, total_pages))

    debut = (page - 1) * par_page
    fin = debut + par_page
    retardataires_page = retardataires_filtres[debut:fin]

    return render_template(
        'comptable/membres_en_retard.html',
        retardataires=retardataires_page,
        mois=mois,
        mois_defaut=mois_defaut,
        nb_retard=nb_retard,
        nb_filtres=nb_filtres,
        recherche=recherche,
        page=page,
        total_pages=total_pages
    )

@app.route('/comptable/cloture-exercice', methods=['GET', 'POST'])
@login_required
@comptable_required
def cloture_exercice():
    bilans = BilanAnnuel.query.order_by(BilanAnnuel.annee.desc()).all()
    annee_courante = datetime.now().year
    
    if request.method == 'POST':
        annee_a_cloturer = request.form.get('annee', type=int)
        if not annee_a_cloturer:
            flash('Veuillez sélectionner une année.', 'danger')
            return redirect(url_for('cloture_exercice'))
        try:
            cloturer_exercice(annee=annee_a_cloturer, clos_par=f"{current_user.prenom} {current_user.nom}")
            rafraichir_dashboard()
            flash(f'Exercice {annee_a_cloturer} clôturé avec succès.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la clôture : {e}', 'danger')
        return redirect(url_for('cloture_exercice'))

    preview = {}
    annee_preview = request.args.get('preview', type=int)
    if annee_preview:
        try:
            preview = get_bilan_annuel(annee_preview)
        except Exception:
            pass

    return render_template('comptable/cloture_exercice.html', bilans=bilans, annee_courante=annee_courante,
                         years_list=list(range(2020, annee_courante + 1)), preview=preview, annee_preview=annee_preview)

# ═════════════════════════════════════════════════════════════════════════════
# GESTION DES ERREURS & LANCEMENT
# ═════════════════════════════════════════════════════════════════════════════
@app.errorhandler(404)
def page_not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(413)
def request_entity_too_large(e):
    flash('Le fichier envoyé est trop volumineux (max 5 Mo).', 'danger')
    return redirect(request.referrer or url_for('mon_compte'))

@app.errorhandler(500)
def internal_server_error(e):
    db.session.rollback()
    return render_template('errors/500.html'), 500

def init_db():
    with app.app_context():
        db.create_all()
        if Membre.query.count() == 0:
            import secrets as secrets_module
            mot_de_passe_genere = secrets_module.token_urlsafe(12)

            admin = Membre(
                nom="Admin",
                prenom="Principal",
                email="admin@mutuelle.com",
                is_admin=True,
                doit_changer_mot_de_passe=True
            )
            admin.set_password(mot_de_passe_genere)
            db.session.add(admin)
            db.session.commit()
            print(f"[OK] Admin créé : admin@mutuelle.com / {mot_de_passe_genere}")
            print("[IMPORTANT] Notez ce mot de passe maintenant, il ne sera plus jamais affiché.")
            
# Route /api/traiter-assistance supprimée — utiliser valider_assistance et refuser_assistance

@app.route('/admin/assistances')
@login_required
def admin_assistances():
    if not (current_user.is_admin or current_user.is_comptable):
        abort(403)
    statut_filtre = request.args.get('statut', '')
    type_filtre   = request.args.get('type', '')

    q = AssistanceEvenement.query.join(
        FamilleMembre, AssistanceEvenement.famille_id == FamilleMembre.id, isouter=True
    )
    if statut_filtre:
        q = q.filter(AssistanceEvenement.statut == statut_filtre)
    else:
        q = q.filter(AssistanceEvenement.statut == 'En attente')
    if type_filtre:
        q = q.filter(AssistanceEvenement.type_evenement == type_filtre)

    demandes = q.order_by(AssistanceEvenement.demande_le.desc()).all()
    return render_template('admin_assistances.html',
                           demandes=demandes,
                           statut_filtre=statut_filtre,
                           type_filtre=type_filtre)


@app.route('/api/enregistrer-paiement', methods=['POST'])
@login_required
@csrf.exempt
def enregistrer_paiement():
    try:
        data = request.get_json()
        
        # 1. Validation des données reçues
        if not all(k in data for k in ('type_paiement', 'mode_paiement', 'montant')):
            return jsonify({'success': False, 'message': 'Données incomplètes'}), 400

        montant = float(data['montant'])
        if montant <= 0:
            return jsonify({'success': False, 'message': 'Montant invalide'}), 400

        # 2. Générer une référence unique pour le suivi
        ref_transaction = f"TXN-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

        # 3. Création de l'objet Cotisation
        nouvelle_cotisation = Cotisation(
            membre_id=current_user.id,
            montant=montant,
            date_paiement=datetime.strptime(data.get('date', datetime.now().strftime('%Y-%m-%d')), '%Y-%m-%d').date(),
            type_cotisation=data['type_paiement'],     # ex: 'mensuel' — colonne affichée dans "Type Cotisation"
            type_paiement=data['type_paiement'],       # conservé pour compatibilité avec le reste du code
            mode_paiement=data['mode_paiement'],        # ex: 'wave', 'orange'
            telephone=data.get('telephone', ''),        # ex: '0707070707'
            reference_transaction=ref_transaction,
            statut='en_attente' # Passe à 'valide' après confirmation API ou validation manuelle
        )

        # 4. Sauvegarde en base de données
        db.session.add(nouvelle_cotisation)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Demande de paiement enregistrée avec succès.',
            'reference': ref_transaction
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur DB enregistrement paiement: {e}")
        return jsonify({'success': False, 'message': 'Erreur lors de l\'enregistrement'}), 500
from sqlalchemy import text

def calculer_solde_caisse():
    try:
        # On appelle la fonction et on récupère la première (et unique) ligne
        result = db.session.execute(text("SELECT * FROM fn_solde_caisse()")).fetchone()
        
        if result:
            # On utilise les noms des colonnes définis dans le RETURNS TABLE de la fonction SQL
            entrees = float(result.total_entrees or 0)
            sorties = float(result.total_sorties or 0)
            solde = float(result.solde or 0)
            
            return entrees, sorties, solde
            
        return 0.0, 0.0, 0.0
        
    except Exception as e:
        current_app.logger.error(f"[ERREUR] Erreur calcul solde caisse: {e}")
        db.session.rollback()
        return 0.0, 0.0, 0.0 

@app.route('/parametres/cotisation', methods=['GET', 'POST'])
@login_required
@admin_required
def parametres_cotisation():
    parametres = ParametresMutuelle.get()
    form = ParametresMutuelleForm(obj=parametres)

    if form.validate_on_submit():
        parametres.modifier(
            frequence_cotisation=form.frequence_cotisation.data,
            montant_cotisation_annuel=form.montant_cotisation_annuel.data
        )
        flash('Paramètres de cotisation mis à jour avec succès.', 'success')
        return redirect(url_for('dashboard'))

    return render_template(
        'finance/parametres_cotisation.html',
        form=form,
        parametres=parametres
    )   

@app.route('/rapports/export/excel')
@login_required
@admin_required
def export_rapport_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    annee = request.args.get('annee', datetime.now().year, type=int)
    d = calculer_donnees_rapport(annee)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Résumé"
    entete_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    entete_font = Font(color='FFFFFF', bold=True)

    ws1.append(['Indicateur', 'Montant (FCFA)'])
    for cell in ws1[1]:
        cell.fill, cell.font = entete_fill, entete_font

    for label, valeur in [
        ('Total Recettes', d['total_recettes']), ('Total Dépenses', d['total_depenses']),
        ('Solde Annuel', d['solde_annuel']), ('Cotisations', d['total_cotisations_annee']),
        ('Dons', d['total_dons_annee']), ('Aides versées', d['total_aides_annee']),
        ('Membres actifs', d['nb_membres_actifs']),
    ]:
        ws1.append([label, valeur])

    ws2 = wb.create_sheet("Détail mensuel")
    ws2.append(['Mois', 'Recettes', 'Dépenses', 'Solde'])
    for cell in ws2[1]:
        cell.fill, cell.font = entete_fill, entete_font
    for m in d['details_mensuels']:
        ws2.append([m['nom'], m['total_recettes'], m['total_depenses'], m['solde_mensuel']])

    for ws in (ws1, ws2):
        for col in ws.columns:
            largeur = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(largeur + 2, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'rapport_financier_{annee}.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/rapports/export/pdf')
@login_required
@admin_required
def export_rapport_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO

    annee = request.args.get('annee', datetime.now().year, type=int)
    d = calculer_donnees_rapport(annee)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Rapport Financier — Année {annee}", styles['Title']),
        Spacer(1, 0.5*cm),
    ]

    resume = [
        ['Indicateur', 'Montant (FCFA)'],
        ['Total Recettes', f"{d['total_recettes']:,.0f}"],
        ['Total Dépenses', f"{d['total_depenses']:,.0f}"],
        ['Solde Annuel', f"{d['solde_annuel']:,.0f}"],
        ['Cotisations', f"{d['total_cotisations_annee']:,.0f}"],
        ['Dons', f"{d['total_dons_annee']:,.0f}"],
        ['Aides versées', f"{d['total_aides_annee']:,.0f}"],
        ['Membres actifs', str(d['nb_membres_actifs'])],
    ]
    t1 = Table(resume, colWidths=[8*cm, 6*cm])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    elements += [t1, Spacer(1, 1*cm)]

    entetes_mois = ['Mois', 'Recettes', 'Dépenses', 'Solde']
    data_mois = [entetes_mois] + [
        [m['nom'], f"{m['total_recettes']:,.0f}", f"{m['total_depenses']:,.0f}", f"{m['solde_mensuel']:,.0f}"]
        for m in d['details_mensuels']
    ]
    t2 = Table(data_mois, repeatRows=1)
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(t2)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'rapport_financier_{annee}.pdf', mimetype='application/pdf')

@app.route('/cotisations/importer', methods=['GET', 'POST'])
@login_required
@comptable_required
def importer_cotisations():
    if request.method == 'GET':
        return render_template('finance/importer_cotisations.html')

    fichier = request.files.get('fichier_import')
    if not fichier or not fichier.filename:
        flash('[ATTENTION] Veuillez sélectionner un fichier.', 'danger')
        return redirect(url_for('importer_cotisations'))

    extension = fichier.filename.rsplit('.', 1)[-1].lower()
    if extension not in ('xlsx', 'xls', 'csv'):
        flash('[ATTENTION] Formats acceptés : .xlsx, .xls, .csv', 'danger')
        return redirect(url_for('importer_cotisations'))

    try:
        if extension == 'csv':
            df = pd.read_csv(fichier, dtype=str, keep_default_na=False)
        else:
            df = pd.read_excel(fichier, dtype=str, keep_default_na=False)
    except Exception as e:
        flash(f'[ERREUR] Fichier illisible : {e}', 'danger')
        return redirect(url_for('importer_cotisations'))

    df.columns = [str(c).strip().lower() for c in df.columns]

    colonnes_obligatoires = ['email', 'montant', 'date_paiement']
    colonnes_manquantes = [c for c in colonnes_obligatoires if c not in df.columns]
    if colonnes_manquantes:
        flash(
            f"[ERREUR] Colonnes obligatoires manquantes : {', '.join(colonnes_manquantes)}. "
            f"Téléchargez le modèle pour le bon format.",
            'danger'
        )
        return redirect(url_for('importer_cotisations'))

    STATUTS_VALIDES = ('en_attente', 'Paye')
    lignes_ok = []
    lignes_erreur = []

    for idx, row in df.iterrows():
        numero_ligne = idx + 2

        email = str(row.get('email', '')).strip().lower()
        if not email:
            lignes_erreur.append((numero_ligne, "email manquant"))
            continue

        membre = Membre.query.filter_by(email=email).first()
        if not membre:
            lignes_erreur.append((numero_ligne, f"aucun membre trouvé avec l'email : {email}"))
            continue

        montant_brut = str(row.get('montant', '')).strip().replace(',', '.')
        try:
            montant = float(montant_brut)
            if montant <= 0:
                raise ValueError
        except ValueError:
            lignes_erreur.append((numero_ligne, f"montant invalide : {montant_brut}"))
            continue

        date_brute = str(row.get('date_paiement', '')).strip().split(' ')[0]
        date_paiement = None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                date_paiement = datetime.strptime(date_brute, fmt)
                break
            except ValueError:
                continue
        if not date_paiement:
            lignes_erreur.append((numero_ligne, f"date_paiement illisible ({date_brute}), format attendu AAAA-MM-JJ"))
            continue

        type_cotisation = str(row.get('type_cotisation', '')).strip() or 'Mensuelle'

        statut_brut = str(row.get('statut', '')).strip()
        statut = statut_brut if statut_brut in STATUTS_VALIDES else 'en_attente'

        mode_paiement = str(row.get('mode_paiement', '')).strip() or None
        reference_transaction = str(row.get('reference_transaction', '')).strip() or None

        try:
            cotisation = Cotisation(
                membre_id=membre.id,
                type_cotisation=type_cotisation,
                montant=montant,
                date_paiement=date_paiement,
                statut=statut,
                mode_paiement=mode_paiement,
                reference_transaction=reference_transaction
            )
            db.session.add(cotisation)
            db.session.commit()
            lignes_ok.append((numero_ligne, email, montant, date_paiement.strftime('%d/%m/%Y')))
        except Exception as e:
            db.session.rollback()
            lignes_erreur.append((numero_ligne, f"erreur enregistrement : {e}"))

    if lignes_ok:
        flash(f"[OK] {len(lignes_ok)} cotisation(s) importée(s) avec succès.", 'success')
    if lignes_erreur:
        flash(f"[ATTENTION] {len(lignes_erreur)} ligne(s) en erreur — voir le détail ci-dessous.", 'warning')

    return render_template(
        'finance/importer_cotisations.html',
        lignes_ok=lignes_ok,
        lignes_erreur=lignes_erreur,
        termine=True
    )


@app.route('/cotisations/modele')
@login_required
@comptable_required
def telecharger_modele_import_cotisations():
    df = pd.DataFrame([{
        'email': 'membre@exemple.com',
        'montant': 5000,
        'date_paiement': '2026-07-09',
        'type_cotisation': 'Mensuelle',
        'statut': 'en_attente',
        'mode_paiement': 'Espèces',
        'reference_transaction': ''
    }])
    buffer = BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=True,
        download_name='modele_import_cotisations.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__': 
    init_db()
    app.run(debug=True)

